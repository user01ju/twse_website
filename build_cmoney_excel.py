"""Build Excel from scraped CMoney category data."""
import json
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

INPUT  = r'C:\Users\Rabbit\claude_code\twse_website\cmoney_raw.json'
OUTPUT = r'C:\Users\Rabbit\claude_code\twse_website\cmoney_categories.xlsx'

HEADER_FONT  = Font(name="Arial", bold=True, color="FFFFFF", size=11)
HEADER_FILL  = PatternFill("solid", start_color="1F4E79")
HEADER_ALIGN = Alignment(horizontal="center", vertical="center")
ALT_FILL     = PatternFill("solid", start_color="EBF3FB")

THIN  = Side(style="thin",   color="C0C0C0")
THICK = Side(style="medium", color="1F4E79")
CELL_BORDER   = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
HEADER_BORDER = Border(left=THICK, right=THICK, top=THICK, bottom=THICK)

PARENT_COLORS = {
    "傳產":   "FFF2CC",
    "電子上游":"DEEAF1",
    "電子中游":"E2EFDA",
    "電子下游":"FCE4D6",
    "軟體":   "EAD1DC",
    "金融":   "D9D9D9",
}


def hdr(ws, row, col, val, extra_font=None):
    c = ws.cell(row=row, column=col, value=val)
    c.font  = extra_font or HEADER_FONT
    c.fill  = HEADER_FILL
    c.alignment = HEADER_ALIGN
    c.border = HEADER_BORDER
    return c


def data_cell(ws, row, col, val, fill=None, bold=False, align="left", num_fmt=None):
    c = ws.cell(row=row, column=col, value=val)
    c.font   = Font(name="Arial", size=10, bold=bold)
    c.fill   = fill or PatternFill()
    c.alignment = Alignment(horizontal=align, vertical="center")
    c.border = CELL_BORDER
    if num_fmt:
        c.number_format = num_fmt
    return c


def build(data):
    wb = Workbook()

    # ── Sheet 1: Flat list ─────────────────────────────────
    ws1 = wb.active
    ws1.title = "類股個股清單"
    ws1.row_dimensions[1].height = 22

    for col, h in enumerate(["大類", "子類股", "類股代碼", "股票代碼", "股票名稱"], 1):
        hdr(ws1, 1, col, h)

    row = 2
    for cat in data:
        parent = cat["parent"]
        pcolor = PatternFill("solid", start_color=PARENT_COLORS.get(parent, "FFFFFF"))
        for sub in cat["subcats"]:
            for i, stock in enumerate(sub["stocks"]):
                fill = pcolor if i % 2 == 0 else ALT_FILL
                data_cell(ws1, row, 1, parent,      fill=pcolor, bold=True, align="center")
                data_cell(ws1, row, 2, sub["name"], fill=fill)
                data_cell(ws1, row, 3, sub["code"], fill=fill, align="center")
                data_cell(ws1, row, 4, stock["id"], fill=fill, align="center")
                data_cell(ws1, row, 5, stock["name"], fill=fill)
                row += 1

    ws1.column_dimensions["A"].width = 12
    ws1.column_dimensions["B"].width = 18
    ws1.column_dimensions["C"].width = 12
    ws1.column_dimensions["D"].width = 10
    ws1.column_dimensions["E"].width = 20
    ws1.freeze_panes = "A2"
    ws1.auto_filter.ref = f"A1:E{row-1}"

    # ── Sheet 2: Summary per sub-category ──────────────────
    ws2 = wb.create_sheet("類股總覽")
    ws2.row_dimensions[1].height = 22

    for col, h in enumerate(["大類", "子類股", "類股代碼", "股票數", "股票代碼列表"], 1):
        hdr(ws2, 1, col, h)

    row2 = 2
    for cat in data:
        parent = cat["parent"]
        pcolor = PatternFill("solid", start_color=PARENT_COLORS.get(parent, "FFFFFF"))
        for i, sub in enumerate(cat["subcats"]):
            fill = pcolor if i % 2 == 0 else ALT_FILL
            codes = "、".join(s["id"] + " " + s["name"] for s in sub["stocks"])
            data_cell(ws2, row2, 1, parent,          fill=pcolor, bold=True, align="center")
            data_cell(ws2, row2, 2, sub["name"],     fill=fill)
            data_cell(ws2, row2, 3, sub["code"],     fill=fill, align="center")
            data_cell(ws2, row2, 4, len(sub["stocks"]), fill=fill, align="center")
            data_cell(ws2, row2, 5, codes,           fill=fill)
            row2 += 1

    ws2.column_dimensions["A"].width = 12
    ws2.column_dimensions["B"].width = 18
    ws2.column_dimensions["C"].width = 12
    ws2.column_dimensions["D"].width = 8
    ws2.column_dimensions["E"].width = 90
    ws2.freeze_panes = "A2"
    ws2.auto_filter.ref = f"A1:D{row2-1}"

    # ── Sheet 3: By parent category ─────────────────────────
    ws3 = wb.create_sheet("大類統計")
    ws3.row_dimensions[1].height = 22

    for col, h in enumerate(["大類", "子類股數", "總股票數"], 1):
        hdr(ws3, 1, col, h)

    row3 = 2
    for cat in data:
        parent    = cat["parent"]
        n_sub     = len(cat["subcats"])
        n_stocks  = sum(len(s["stocks"]) for s in cat["subcats"])
        pcolor    = PatternFill("solid", start_color=PARENT_COLORS.get(parent, "FFFFFF"))
        data_cell(ws3, row3, 1, parent,   fill=pcolor, bold=True, align="center")
        data_cell(ws3, row3, 2, n_sub,    fill=pcolor, align="center")
        data_cell(ws3, row3, 3, n_stocks, fill=pcolor, align="center")
        row3 += 1

    ws3.column_dimensions["A"].width = 14
    ws3.column_dimensions["B"].width = 12
    ws3.column_dimensions["C"].width = 12
    ws3.freeze_panes = "A2"

    wb.save(OUTPUT)
    print(f"Saved: {OUTPUT}")


def main():
    with open(INPUT, encoding="utf-8") as f:
        raw = json.load(f)

    # Reshape: raw is [{parent, name, code, stocks}] flat list
    # Group by parent
    parents_order = []
    by_parent = {}
    for item in raw:
        p = item["parent"]
        if p not in by_parent:
            by_parent[p] = []
            parents_order.append(p)
        by_parent[p].append({
            "name":   item["name"],
            "code":   item["code"],
            "stocks": item["stocks"],
        })

    data = [{"parent": p, "subcats": by_parent[p]} for p in parents_order]

    total = sum(len(sub["stocks"]) for cat in data for sub in cat["subcats"])
    print(f"Building Excel: {len(data)} parent categories, {sum(len(c['subcats']) for c in data)} sub-cats, {total} stocks")
    build(data)


if __name__ == "__main__":
    main()
