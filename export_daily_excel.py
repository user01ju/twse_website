"""Export 漲幅前100 + 外資買超 from latest TWSE data to Excel."""
import sys
import os
sys.path.insert(0, os.path.dirname(__file__))

from concurrent.futures import ThreadPoolExecutor, as_completed
from datetime import date, datetime

from config import BASE_DIR
from fetcher import twse_client, tpex_client
from fetcher.market_calendar import roc_to_date
from processor import movers, foreign_trades
from processor.utils import parse_num

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

OUTPUT = str(BASE_DIR / "daily_export.xlsx")

# ── Style constants ─────────────────────────────────────────
HDR_FONT  = Font(name="Arial", bold=True, color="FFFFFF", size=11)
HDR_FILL  = PatternFill("solid", start_color="1F4E79")
HDR_ALIGN = Alignment(horizontal="center", vertical="center", wrap_text=True)
POS_FONT  = Font(name="Arial", color="C00000", size=10)  # red = up
NEG_FONT  = Font(name="Arial", color="375623", size=10)  # green = down
NRM_FONT  = Font(name="Arial", size=10)
THIN  = Side(style="thin",   color="BFBFBF")
THICK = Side(style="medium", color="1F4E79")
CELL_BDR  = Border(left=THIN,  right=THIN,  top=THIN,  bottom=THIN)
HDR_BDR   = Border(left=THICK, right=THICK, top=THICK, bottom=THICK)
ALT_FILL  = PatternFill("solid", start_color="EBF3FB")
POS_FILL  = PatternFill("solid", start_color="FDECEA")
NEG_FILL  = PatternFill("solid", start_color="E9F5EC")
BUY_FILL  = PatternFill("solid", start_color="DEEAF1")
BUY_ALT   = PatternFill("solid", start_color="F2F7FB")


def hdr(ws, row, col, val):
    c = ws.cell(row=row, column=col, value=val)
    c.font = HDR_FONT; c.fill = HDR_FILL
    c.alignment = HDR_ALIGN; c.border = HDR_BDR
    return c


def cell(ws, row, col, val, fill=None, font=None, align="center", fmt=None):
    c = ws.cell(row=row, column=col, value=val)
    c.font   = font or NRM_FONT
    c.fill   = fill or PatternFill()
    c.alignment = Alignment(horizontal=align, vertical="center")
    c.border = CELL_BDR
    if fmt:
        c.number_format = fmt
    return c


# ── Fetch ───────────────────────────────────────────────────
def fetch_all():
    today = date.today().strftime("%Y%m%d")
    tasks = {
        "twse_stocks": (twse_client.fetch_stock_day_all, [today]),
        "tpex_daily":  (tpex_client.fetch_daily_quotes,  []),
        "twse_t86":    (twse_client.fetch_t86,            []),
        "tpex_qfii":   (tpex_client.fetch_3insti_qfii,   []),
        "twse_index":  (twse_client.fetch_taiex_index,   [today]),
    }
    raw = {}
    with ThreadPoolExecutor(max_workers=5) as pool:
        fmap = {pool.submit(fn, *args): k for k, (fn, args) in tasks.items()}
        for f in as_completed(fmap, timeout=60):
            k = fmap[f]
            try:    raw[k] = f.result(timeout=30)
            except: raw[k] = None
    return raw


def detect_date(twse_stocks, twse_index):
    for src in (twse_stocks, twse_index):
        if src:
            raw_date = src[0].get("Date") or src[0].get("日期")
            if raw_date:
                try:
                    return roc_to_date(str(raw_date).strip())
                except Exception:
                    pass
    return date.today()


def build_price_lookup(stocks, code_field, price_field):
    result = {}
    for s in stocks or []:
        try:
            code = str(s.get(code_field, "")).strip()
            val  = str(s.get(price_field, "0")).replace(",", "").strip()
            result[code] = float(val) if val not in ("---", "--", "") else 0.0
        except: pass
    return result


# ── Sheet: 漲幅前100 ────────────────────────────────────────
def write_movers(wb, gainers, date_str):
    ws = wb.create_sheet("漲幅前100")
    ws.row_dimensions[1].height = 28
    ws.row_dimensions[2].height = 18

    ws.merge_cells("A1:H1")
    title = ws.cell(row=1, column=1, value=f"漲幅前100  |  資料日期：{date_str}")
    title.font = Font(name="Arial", bold=True, size=13, color="1F4E79")
    title.alignment = Alignment(horizontal="left", vertical="center")

    hdrs = ["排名", "代號", "名稱", "市場", "收盤價", "漲跌", "漲跌幅%", "成交量(張)"]
    for c, h in enumerate(hdrs, 1):
        hdr(ws, 2, c, h)

    for i, s in enumerate(gainers, 1):
        r = i + 2
        is_alt = i % 2 == 0
        bg = ALT_FILL if is_alt else PatternFill()
        pct = s["change_pct"]
        fill = POS_FILL if pct > 0 else (NEG_FILL if pct < 0 else bg)
        pfont = POS_FONT if pct > 0 else (NEG_FONT if pct < 0 else NRM_FONT)

        cell(ws, r, 1, i,                fill=fill, align="center")
        cell(ws, r, 2, s["code"],        fill=fill, align="center")
        cell(ws, r, 3, s["name"],        fill=fill, align="left", font=NRM_FONT)
        cell(ws, r, 4, s["market"],      fill=fill, align="center")
        cell(ws, r, 5, s["close"],       fill=fill, font=pfont, fmt="#,##0.00")
        c6 = cell(ws, r, 6, s["change"],  fill=fill, font=pfont, fmt="+#,##0.00;-#,##0.00;-")
        cell(ws, r, 7, s["change_pct"]/100, fill=fill, font=pfont, fmt="+0.00%;-0.00%;-")
        cell(ws, r, 8, s["volume_zhang"], fill=fill, fmt="#,##0")

    ws.column_dimensions["A"].width = 6
    ws.column_dimensions["B"].width = 8
    ws.column_dimensions["C"].width = 16
    ws.column_dimensions["D"].width = 7
    ws.column_dimensions["E"].width = 10
    ws.column_dimensions["F"].width = 10
    ws.column_dimensions["G"].width = 10
    ws.column_dimensions["H"].width = 12
    ws.freeze_panes = "A3"
    ws.auto_filter.ref = f"A2:H{len(gainers)+2}"


# ── Sheet: 外資買超 ─────────────────────────────────────────
def write_foreign(wb, buy_super, date_str):
    ws = wb.create_sheet("外資買超")
    ws.row_dimensions[1].height = 28
    ws.row_dimensions[2].height = 18

    ws.merge_cells("A1:G1")
    title = ws.cell(row=1, column=1, value=f"外資買超  |  資料日期：{date_str}")
    title.font = Font(name="Arial", bold=True, size=13, color="1F4E79")
    title.alignment = Alignment(horizontal="left", vertical="center")

    hdrs = ["排名", "代號", "名稱", "市場", "買超(張)", "賣超(張)", "淨買超(張)", "淨買超(億)"]
    for c, h in enumerate(hdrs, 1):
        hdr(ws, 2, c, h)

    for i, s in enumerate(buy_super, 1):
        r = i + 2
        fill = BUY_FILL if i % 2 == 1 else BUY_ALT

        cell(ws, r, 1, i,                   fill=fill, align="center")
        cell(ws, r, 2, s["code"],            fill=fill, align="center")
        cell(ws, r, 3, s["name"],            fill=fill, align="left", font=NRM_FONT)
        cell(ws, r, 4, s["market"],          fill=fill, align="center")
        cell(ws, r, 5, s["buy_zhang"],       fill=fill, font=POS_FONT, fmt="#,##0")
        cell(ws, r, 6, s["sell_zhang"],      fill=fill, font=NEG_FONT, fmt="#,##0")
        cell(ws, r, 7, s["net_zhang"],       fill=fill, font=POS_FONT, fmt="#,##0")
        cell(ws, r, 8, round(s["net_yi"],2), fill=fill, font=POS_FONT, fmt="#,##0.00")

    ws.column_dimensions["A"].width = 6
    ws.column_dimensions["B"].width = 8
    ws.column_dimensions["C"].width = 16
    ws.column_dimensions["D"].width = 7
    ws.column_dimensions["E"].width = 12
    ws.column_dimensions["F"].width = 12
    ws.column_dimensions["G"].width = 12
    ws.column_dimensions["H"].width = 12
    ws.freeze_panes = "A3"
    ws.auto_filter.ref = f"A2:H{len(buy_super)+2}"


def main():
    print("Fetching data...")
    raw = fetch_all()

    twse_stocks = raw.get("twse_stocks") or []
    tpex_daily  = raw.get("tpex_daily")  or []
    twse_t86    = raw.get("twse_t86")    or {}
    tpex_qfii   = raw.get("tpex_qfii")  or []

    data_date = detect_date(twse_stocks, raw.get("twse_index"))
    date_str  = data_date.isoformat()
    print(f"Data date: {date_str}")

    twse_prices = build_price_lookup(twse_stocks, "Code", "ClosingPrice")
    tpex_prices = build_price_lookup(tpex_daily,  "SecuritiesCompanyCode", "Close")

    # 漲幅前100
    mv = movers.build(twse_stocks, tpex_daily)
    gainers = mv["gainers"][:100]
    print(f"Gainers: {len(gainers)}")

    # 外資買超
    ft = foreign_trades.build(twse_t86, tpex_qfii, twse_prices, tpex_prices)
    buy_super = ft["buy_super"]  # already sorted, up to 100
    print(f"Foreign buy: {len(buy_super)}")

    # Build Excel
    wb = Workbook()
    wb.remove(wb.active)  # remove default sheet
    write_movers(wb,  gainers,   date_str)
    write_foreign(wb, buy_super, date_str)

    wb.save(OUTPUT)
    print(f"\nSaved: {OUTPUT}")
    print(f"  漲幅前100: {len(gainers)} stocks")
    print(f"  外資買超:  {len(buy_super)} stocks")


if __name__ == "__main__":
    main()
