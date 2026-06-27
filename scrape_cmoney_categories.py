"""
Scrape CMoney forum category pages to extract all sector categories and their stocks.
Uses the __NUXT__ SSR data embedded in each page for clean structured extraction.
"""
import json
import re
import time
import requests
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

HEADERS = {
    "User-Agent": (
        "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
        "AppleWebKit/537.36 (KHTML, like Gecko) "
        "Chrome/124.0.0.0 Safari/537.36"
    ),
    "Accept-Language": "zh-TW,zh;q=0.9",
    "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8",
}

BASE_URL = "https://www.cmoney.tw"
OVERVIEW_URL = f"{BASE_URL}/forum/category"


def fetch_nuxt_data(url: str) -> dict | None:
    """Fetch a CMoney page and extract __NUXT__ SSR data."""
    try:
        r = requests.get(url, headers=HEADERS, timeout=15)
        r.raise_for_status()
    except Exception as e:
        print(f"  ERROR fetching {url}: {e}")
        return None

    html = r.text
    # Extract window.__NUXT__=(...) from script tag
    m = re.search(r'window\.__NUXT__\s*=\s*(\{.*?\});?\s*</script>', html, re.DOTALL)
    if not m:
        # Try alternate pattern
        m = re.search(r'window\.__NUXT__\s*=\s*(\(.*?\));?\s*</script>', html, re.DOTALL)
    if not m:
        print(f"  No __NUXT__ data found at {url}")
        return None

    try:
        data_str = m.group(1)
        # Some nuxt versions wrap in ()
        if data_str.startswith('('):
            data_str = data_str[1:]
            if data_str.endswith(')'):
                data_str = data_str[:-1]
        return json.loads(data_str)
    except json.JSONDecodeError as e:
        print(f"  JSON parse error at {url}: {e}")
        return None


def get_all_categories() -> list[dict]:
    """Scrape the overview page for all parent/sub categories."""
    print("Fetching overview page...")
    nuxt = fetch_nuxt_data(OVERVIEW_URL)
    if not nuxt:
        # Fallback: use the hardcoded list we already observed from DOM
        print("  Using fallback category list...")
        return get_fallback_categories()

    # Try to find category list in nuxt data
    data0 = (nuxt.get("data") or [{}])[0]
    cat_data = data0.get("categoryList") or data0.get("categories") or None

    if not cat_data:
        print("  No categoryList in SSR data, using fallback...")
        return get_fallback_categories()

    return cat_data


def get_fallback_categories() -> list[dict]:
    """Hardcoded categories extracted from DOM earlier."""
    return [
        {"parent": "傳產", "subcats": [
            {"name": "水泥",     "code": "C11010"},
            {"name": "食品",     "code": "C12010"},
            {"name": "塑膠",     "code": "C13010"},
            {"name": "紡織纖維", "code": "C14010"},
            {"name": "電機",     "code": "C15010"},
            {"name": "電線電纜", "code": "C16010"},
            {"name": "化學工業", "code": "C17010"},
            {"name": "生技",     "code": "C17020"},
            {"name": "玻璃陶瓷", "code": "C18010"},
            {"name": "紙業",     "code": "C19010"},
            {"name": "鋼鐵",     "code": "C20010"},
            {"name": "橡膠",     "code": "C21010"},
            {"name": "汽車",     "code": "C22010"},
            {"name": "汽車零組件","code": "C22020"},
            {"name": "營建",     "code": "C25010"},
            {"name": "航運",     "code": "C26010"},
            {"name": "觀光",     "code": "C27010"},
            {"name": "金融",     "code": "C28010"},
            {"name": "貿易百貨", "code": "C29010"},
            {"name": "油電燃氣", "code": "C23010"},
            {"name": "綜合",     "code": "C99010"},
        ]},
        {"parent": "電子", "subcats": [
            {"name": "半導體",   "code": "C31010"},
            {"name": "電腦週邊", "code": "C32010"},
            {"name": "光電",     "code": "C33010"},
            {"name": "通信網路", "code": "C34010"},
            {"name": "電子零組件","code": "C35010"},
            {"name": "電子通路", "code": "C36010"},
            {"name": "資訊服務", "code": "C37010"},
            {"name": "其他電子", "code": "C39010"},
            {"name": "電機機械", "code": "C15010"},
        ]},
        {"parent": "上櫃", "subcats": [
            {"name": "上櫃半導體","code": "CB1010"},
            {"name": "上櫃電腦週邊","code": "CB2010"},
            {"name": "上櫃光電", "code": "CB3010"},
            {"name": "上櫃通信網路","code": "CB4010"},
            {"name": "上櫃電子零組件","code": "CB5010"},
            {"name": "上櫃電子通路","code": "CB6010"},
            {"name": "上櫃資訊服務","code": "CB7010"},
            {"name": "上櫃其他電子","code": "CB9010"},
            {"name": "上櫃食品",  "code": "CA2010"},
            {"name": "上櫃化學",  "code": "CA7010"},
            {"name": "上櫃生技",  "code": "CA7020"},
            {"name": "上櫃鋼鐵",  "code": "CA0010"},
            {"name": "上櫃營建",  "code": "CA5010"},
            {"name": "上櫃航運",  "code": "CA6010"},
            {"name": "上櫃觀光",  "code": "CA7030"},
            {"name": "上櫃金融",  "code": "CA8010"},
            {"name": "上櫃綜合",  "code": "CA9010"},
        ]},
    ]


def get_stocks_for_category(code: str) -> list[dict]:
    """Fetch a category page and return its stock list."""
    url = f"{BASE_URL}/forum/category/{code}"
    nuxt = fetch_nuxt_data(url)
    if not nuxt:
        return []

    data0 = (nuxt.get("data") or [{}])[0]
    stock_list = data0.get("stockList") or []

    result = []
    for s in stock_list:
        sid = str(s.get("stockId", "")).strip()
        # Only keep 4-digit numeric codes
        if not (len(sid) == 4 and sid.isdigit()):
            continue
        result.append({
            "stock_id":   sid,
            "stock_name": s.get("stockName", ""),
        })
    return result


def build_excel(all_data: list[dict], output_path: str) -> None:
    """Build Excel with two sheets: flat list + summary."""
    wb = Workbook()

    # ── Sheet 1: Flat list ──────────────────────────────────
    ws1 = wb.active
    ws1.title = "類股個股清單"

    header_font = Font(name="Arial", bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", start_color="1F4E79")
    thin = Side(style="thin", color="CCCCCC")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    headers = ["大類", "子類股", "類股代碼", "股票代碼", "股票名稱"]
    for col, h in enumerate(headers, 1):
        cell = ws1.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")
        cell.border = border

    row = 2
    for cat in all_data:
        for sub in cat["subcats"]:
            for stock in sub["stocks"]:
                ws1.cell(row=row, column=1, value=cat["parent"])
                ws1.cell(row=row, column=2, value=sub["name"])
                ws1.cell(row=row, column=3, value=sub["code"])
                ws1.cell(row=row, column=4, value=stock["stock_id"])
                ws1.cell(row=row, column=5, value=stock["stock_name"])
                for col in range(1, 6):
                    ws1.cell(row=row, column=col).border = border
                    ws1.cell(row=row, column=col).font = Font(name="Arial")
                row += 1

    ws1.column_dimensions["A"].width = 12
    ws1.column_dimensions["B"].width = 16
    ws1.column_dimensions["C"].width = 12
    ws1.column_dimensions["D"].width = 10
    ws1.column_dimensions["E"].width = 16
    ws1.freeze_panes = "A2"

    # ── Sheet 2: Summary by sub-category ───────────────────
    ws2 = wb.create_sheet("類股總覽")

    for col, h in enumerate(["大類", "子類股", "類股代碼", "股票數量", "股票代碼列表"], 1):
        cell = ws2.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")
        cell.border = border

    row2 = 2
    for cat in all_data:
        for sub in cat["subcats"]:
            codes = ", ".join(s["stock_id"] for s in sub["stocks"])
            ws2.cell(row=row2, column=1, value=cat["parent"])
            ws2.cell(row=row2, column=2, value=sub["name"])
            ws2.cell(row=row2, column=3, value=sub["code"])
            ws2.cell(row=row2, column=4, value=len(sub["stocks"]))
            ws2.cell(row=row2, column=5, value=codes)
            for col in range(1, 6):
                ws2.cell(row=row2, column=col).border = border
                ws2.cell(row=row2, column=col).font = Font(name="Arial")
            row2 += 1

    ws2.column_dimensions["A"].width = 12
    ws2.column_dimensions["B"].width = 16
    ws2.column_dimensions["C"].width = 12
    ws2.column_dimensions["D"].width = 10
    ws2.column_dimensions["E"].width = 60
    ws2.freeze_panes = "A2"

    wb.save(output_path)
    print(f"\nSaved: {output_path}")


def main():
    categories = get_fallback_categories()

    # Count total subcategories
    total = sum(len(c["subcats"]) for c in categories)
    print(f"Total sub-categories to scrape: {total}")

    all_data = []
    done = 0

    for cat in categories:
        parent = cat["parent"]
        enriched_subcats = []

        for sub in cat["subcats"]:
            done += 1
            code = sub["code"]
            name = sub["name"]
            print(f"[{done}/{total}] {parent} > {name} ({code})...", end=" ")

            stocks = get_stocks_for_category(code)
            print(f"{len(stocks)} stocks")

            enriched_subcats.append({
                "name":   name,
                "code":   code,
                "stocks": stocks,
            })
            time.sleep(0.4)  # polite delay

        all_data.append({"parent": parent, "subcats": enriched_subcats})

    # Summary
    total_stocks = sum(
        len(sub["stocks"])
        for cat in all_data
        for sub in cat["subcats"]
    )
    print(f"\nTotal stocks collected: {total_stocks}")

    output = r"C:\Users\Rabbit\claude_code\twse_website\cmoney_categories.xlsx"
    build_excel(all_data, output)


if __name__ == "__main__":
    main()
