"""Fetch all API endpoints and export to Excel with one sheet per API."""
import sys
import os
sys.path.insert(0, os.path.dirname(__file__))

from datetime import date
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

from fetcher import twse_client, tpex_client

TODAY = date.today()
TODAY_STR = TODAY.strftime("%Y%m%d")
# Also try previous trading day for legacy endpoints
PREV_STR = "20260507"

OUTPUT_FILE = f"api_data_{TODAY_STR}.xlsx"


def norm_rows(data) -> tuple[list[str], list[list]]:
    """Normalize API response to (headers, rows). Handles list[dict] and dict with fields+data."""
    if isinstance(data, dict):
        fields = data.get("fields", [])
        rows   = data.get("data", [])
        return fields, rows
    elif isinstance(data, list) and data:
        headers = list(data[0].keys())
        rows    = [[r.get(h, "") for h in headers] for r in data]
        return headers, rows
    return [], []


def write_sheet(wb: openpyxl.Workbook, sheet_name: str, headers: list[str], rows: list[list], limit: int = 1000):
    ws = wb.create_sheet(title=sheet_name)

    header_font  = Font(bold=True, color="FFFFFF")
    header_fill  = PatternFill("solid", start_color="2F5496")
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)

    # Write headers
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=str(h))
        cell.font   = header_font
        cell.fill   = header_fill
        cell.alignment = header_align

    # Write data rows (up to limit)
    for row_idx, row in enumerate(rows[:limit], 2):
        for col_idx, val in enumerate(row, 1):
            ws.cell(row=row_idx, column=col_idx, value=val)

    # Auto column widths
    for col_idx, h in enumerate(headers, 1):
        col_letter = get_column_letter(col_idx)
        max_len = len(str(h))
        for row in rows[:50]:
            if col_idx - 1 < len(row):
                max_len = max(max_len, len(str(row[col_idx - 1])))
        ws.column_dimensions[col_letter].width = min(max_len + 2, 30)

    # Freeze first row
    ws.freeze_panes = "A2"

    ws.row_dimensions[1].height = 30
    print(f"  [{sheet_name}] {len(rows[:limit])} rows × {len(headers)} cols")


def fetch_with_fallback(fn, *args):
    try:
        return fn(*args)
    except Exception as e:
        print(f"  WARNING: {fn.__name__}{args} failed: {e}")
        return None


wb = openpyxl.Workbook()
wb.remove(wb.active)  # Remove default sheet

# ── 1. TWSE MI_INDEX ──────────────────────────────────────────────────────────
print("Fetching TWSE MI_INDEX...")
data = fetch_with_fallback(twse_client.fetch_taiex_index)
if data:
    h, r = norm_rows(data)
    write_sheet(wb, "TWSE_MI_INDEX", h, r)

# ── 2. TWSE STOCK_DAY_ALL ─────────────────────────────────────────────────────
print("Fetching TWSE STOCK_DAY_ALL...")
data = fetch_with_fallback(twse_client.fetch_stock_day_all)
if data:
    h, r = norm_rows(data)
    write_sheet(wb, "TWSE_STOCK_DAY_ALL", h, r)

# ── 3. TWSE BFI82U (三大法人彙總) ───────────────────────────────────────────
print("Fetching TWSE BFI82U...")
data = None
for ds in [TODAY_STR, PREV_STR]:
    data = fetch_with_fallback(twse_client.fetch_bfi82u, ds)
    if data and data.get("data"):
        print(f"  BFI82U data date: {ds}")
        break
if data:
    h, r = norm_rows(data)
    write_sheet(wb, "TWSE_BFI82U", h, r)

# ── 4. TWSE T86 (三大法人 by stock) ─────────────────────────────────────────
print("Fetching TWSE T86...")
data = None
for ds in [TODAY_STR, PREV_STR]:
    data = fetch_with_fallback(twse_client.fetch_t86, ds)
    if data and data.get("data"):
        print(f"  T86 data date: {ds}")
        break
if data:
    h, r = norm_rows(data)
    write_sheet(wb, "TWSE_T86", h, r)

# ── 5. TPEX Daily Quotes ──────────────────────────────────────────────────────
print("Fetching TPEX Daily Quotes...")
data = fetch_with_fallback(tpex_client.fetch_daily_quotes)
if data:
    h, r = norm_rows(data)
    write_sheet(wb, "TPEX_Daily_Quotes", h, r)

# ── 6. TPEX 3insti Summary ────────────────────────────────────────────────────
print("Fetching TPEX 3insti Summary...")
data = fetch_with_fallback(tpex_client.fetch_3insti_summary)
if data:
    h, r = norm_rows(data)
    write_sheet(wb, "TPEX_3insti_Summary", h, r)

# ── 7. TPEX 3insti QFII (外資) ───────────────────────────────────────────────
print("Fetching TPEX 3insti QFII...")
data = fetch_with_fallback(tpex_client.fetch_3insti_qfii)
if data:
    h, r = norm_rows(data)
    write_sheet(wb, "TPEX_3insti_QFII", h, r)

# ── 8. TPEX 3insti Trust (投信) ──────────────────────────────────────────────
print("Fetching TPEX 3insti Trust...")
data = fetch_with_fallback(tpex_client.fetch_3insti_trust)
if data:
    h, r = norm_rows(data)
    write_sheet(wb, "TPEX_3insti_Trust", h, r)

# ── 9. TPEX 3insti All (三大合計) ────────────────────────────────────────────
print("Fetching TPEX 3insti All...")
data = fetch_with_fallback(tpex_client.fetch_3insti_all)
if data:
    h, r = norm_rows(data)
    write_sheet(wb, "TPEX_3insti_All", h, r)

# ── 10. TPEX ESB Quotes (興櫃) ───────────────────────────────────────────────
print("Fetching TPEX ESB Quotes...")
data = fetch_with_fallback(tpex_client.fetch_esb_quotes)
if data:
    h, r = norm_rows(data)
    write_sheet(wb, "TPEX_ESB_Quotes", h, r)

wb.save(OUTPUT_FILE)
print(f"\nSaved: {OUTPUT_FILE}")
